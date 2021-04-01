# -*- coding: utf-8 -*-
# @Time : 2021/1/24 3:47 下午
# @Author : Qi Tian yue
# @Github : Orange-66
# @PROJECT : LAPS 
# @File : tool_file.py
# @Remark : 控制文件的工具类
# -----------------------------
import os
import random
import shutil
import zipfile
import openpyxl
import send2trash
from PyQt5.QtGui import QPixmap
from PyQt5.QtWidgets import QFileDialog, QMessageBox
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from PyQt5.QtCore import QFile, Qt, QIODevice
from shutil import copy, move
from Utils import settings, tool_win, tool_log

# 加载qss文件
from Window.win_index import Win_Index


def load_qss(qss_path):
    file = QFile(qss_path)
    file.open(QFile.ReadOnly)

    # QByteArray
    qt_bytes = file.readAll()
    # QByteArray转换为bytes
    py_bytes = qt_bytes.data()
    # bytes转换为str
    style_sheet = py_bytes.decode("utf-8")

    settings.app.setStyleSheet(style_sheet)


# 压缩文件
def compress_file():
    filename = "Laps.zip"
    zip_item = zipfile.ZipFile(filename, 'w')
    zip_item.write('./EXE', compress_type=zipfile.ZIP_DEFLATED)
    zip_item.close()


# 加载Excel文件
def load_excel(file_path):
    return openpyxl.load_workbook(file_path)


# 得到第一个工作表
def get_worksheet_by_index(work_book, index):
    sheet_names = work_book.get_sheet_names()
    work_sheet = work_book.get_sheet_by_names(sheet_names[index])
    return work_sheet


# 获取指定一行数据
def get_row_by_index(work_sheet, index):
    return work_sheet.row[index]


# 获取最大数据行数
def get_range_row(work_sheet):
    return range(2, work_sheet.get_highest_row())


# 创建一个excel文档
def create_sample(save_dir):
    # 获取excel实例对象
    work_book = openpyxl.Workbook()

    # 设置工作表名
    sheet_name = work_book.get_sheet_names()[0]
    sheet = work_book.get_sheet_by_name(sheet_name)
    sheet.title = settings.excel_sheet_name

    # 设置标题样式
    title_style = Font(size=settings.excel_font_size, bold=settings.excel_bold)

    # 冻结表头
    sheet.freeze_panes = settings.excel_freeze_strategy

    # 填写表头
    for i in range(len(settings.excel_sheet_titles)):
        column = get_column_letter(i + 1)
        # 设置表头样式以及内容
        sheet[column + '1'].font = title_style
        sheet[column + '1'] = settings.excel_sheet_titles[i]
        # 设置表头的宽
        sheet.column_dimensions[column].width = settings.excel_title_width

    # 保存excel表与指定位置
    work_book.save(os.path.join(save_dir, settings.excel_file_name))


# 打开图片文件
def open_image(window):
    image_name_list, _ = \
        QFileDialog.getOpenFileNames(window, "Open Image File", os.getcwd(), "Image(*.jpg *.png *.jpeg)")
    tool_log.debug(image_name_list)
    image_pix_list = []
    for image_name in image_name_list:
        image_pix = QPixmap(image_name)
        image_pix_list.append(image_pix)
    return image_name_list, image_pix_list


# 根据文件路径截取文件名
def get_file_name(file_path):
    file_name = file_path.split(os.sep)[-1]
    tool_log.debug("get_file_name", file_path, file_name)
    return file_name


# 打开图片文件，读取为二进制数据
def open_image_file(image_path):
    file = QFile(image_path)
    if not file.open(QIODevice.ReadOnly):
        tool_log.debug("tool_db - insert_image, 该路径下无此文件，" + image_path)
        QMessageBox.warning(settings.win_index, "错误", "该路径下无此文件," + image_path)
        return
    else:
        image = file.readAll()
        file.close()
        return image


# 剪切文件到文件夹
def move_file(src_path, dst_path):
    make_dir(dst_path)

    move(src_path, dst_path)


# 复制文件到文件夹
def copy_file(src_path, dst_path):
    make_dir(dst_path)

    copy(src_path, dst_path)


# 重命名文件
def rename_file(filename, newFilename):
    temp = filename.split('.')
    temp[0] = newFilename

    return '.'.join(temp)


# 根据路径字符串一层一层创建文件夹
def make_dir(dir_string):
    for i in range(len(dir_string.split("/")) - 1):
        dir_list = []
        for j in range(i + 1):
            dir_list.append(dir_string.split('/')[j])
        dirs = '/'.join(dir_list)
        if not os.path.exists(dirs):
            os.mkdir(dirs)


# 连接路径字符串
def make_path(*args):
    result_list = []
    for i in args:
        result_list.append(str(i))

    return '/'.join(result_list)


# 根据路径获取文件名
def get_filename(file_path):
    return file_path.split('/')[-1]


# 随机赋值文件名
def make_filename():
    # 多个字符中选取指定数量的字符组成新字符串：
    return ''.join(random.sample(
        ['z', 'y', 'x', 'w', 'v', 'u', 't', 's', 'r', 'q', 'p', 'o', 'n', 'm', 'l', 'k', 'j', 'i', 'h', 'g', 'f', 'e',
         'd', 'c', 'b', 'a', '0', '1', '2', '3', '4', '5', '6', '7', '8', '9'], 8))


# 删除指定文件夹
def delete_file(file_dir, delete_mode):
    tool_log.debug('delete_file: ', file_dir, " - ", delete_mode)
    # ever
    if delete_mode == 'e':
        shutil.rmtree(file_dir)
    # trash
    elif delete_mode == 't':
        send2trash.send2trash(file_dir)
